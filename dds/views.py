from datetime import datetime, time
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from .forms import DDSOpCreateForm, DDSOperationForm, DDSArticleForm
from .models import DDSOperation, DDSArticle, Hotel, DDSArticle, DDSOperation, CashMovement, CashRegister
from django.db import IntegrityError
from django.shortcuts import redirect
from .forms import DDSOperationForm
from .cash_services import apply_cash_movement
from django.db.models import Sum, Q, F
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from openpyxl import Workbook
from decimal import Decimal, ROUND_HALF_UP
from django.contrib import messages
from django.shortcuts import redirect
from .forms import CashIncassoForm
from .models import CashIncasso
from django.db import transaction
from django.core.exceptions import ValidationError
from .cash_services import apply_cash_movement, FIELD_MAP
from collections import defaultdict
from django.db.models import Sum, Q, F
from django.db.models.functions import Coalesce, TruncDate
from django.shortcuts import redirect, render, get_object_or_404
from django.db.models.functions import Coalesce, TruncDate
from django.db.models import Sum, Q
from .utils import user_hotels_qs
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from .forms import DDSOperationForm
from .models import DDSArticle,DDSArticle, Hotel, DDSOperation, DDSArticle, CashRegister, CashRegister, DDSOperation, DDSOperation, CashMovement, CashRegister, CashMovement, DDSArticle
from .cash_services import apply_cash_movement, FIELD_MAP
from django.db import transaction
from django.core.exceptions import ValidationError
from .cash_services import apply_cash_movement, FIELD_MAP
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import CashRegister, DDSArticle
from .forms import HotelForm
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce, TruncDate
from collections import OrderedDict
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def _parse_date(d: str):
    try:
        return datetime.strptime(d, "%Y-%m-%d").date()
    except Exception:
        return None



def _day_range(date_obj):
    start = timezone.make_aware(datetime.combine(date_obj, time.min))
    end = timezone.make_aware(datetime.combine(date_obj, time.max))
    return start, end


# ── Excel helpers ────────────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # тёмно-синий
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")   # светло-голубой
_TOTAL_FONT = Font(bold=True, size=10)
_THIN       = Side(style="thin", color="AAAAAA")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NUM_FMT    = '#,##0.00'
_PCT_FMT    = '0.00%'


def _pct(part, total):
    """Доля part/total в виде 0.0–1.0 для Excel-формата '0.00%'."""
    try:
        return float(part) / float(total) if total else 0.0
    except (TypeError, ZeroDivisionError):
        return 0.0


def _xl_hdr(ws, row, headers):
    """Запись + стилизация строки заголовка."""
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20


def _xl_row(ws, row, values, money_cols=(), pct_cols=()):
    """Запись + стилизация строки данных."""
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = _BORDER
        if col in money_cols:
            c.number_format = _NUM_FMT
            c.alignment = Alignment(horizontal="right")
        elif col in pct_cols:
            c.number_format = _PCT_FMT
            c.alignment = Alignment(horizontal="right")


def _xl_total_row(ws, row, values, money_cols=(), pct_cols=()):
    """Запись + стилизация итоговой строки."""
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _TOTAL_FILL
        c.font = _TOTAL_FONT
        c.border = _BORDER
        if col in money_cols:
            c.number_format = _NUM_FMT
            c.alignment = Alignment(horizontal="right")
        elif col in pct_cols:
            c.number_format = _PCT_FMT
            c.alignment = Alignment(horizontal="right")


def _xl_widths(ws, pairs):
    """Установка ширины колонок: pairs = [('A', 20), ('B', 14), ...]."""
    for letter, w in pairs:
        ws.column_dimensions[letter].width = w
# ────────────────────────────────────────────────────────────────────────────


@login_required
def hotel_detail_export_excel(request, pk):
    hotels_qs = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels_qs, pk=pk)

    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    ops = (DDSOperation.objects
           .select_related("article")
           .filter(hotel=hotel, is_voided=False))

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    rooms_q = (
        Q(source__iexact="rooms") |
        Q(source__icontains="room") |
        Q(article__name__icontains="номер") |
        Q(article__name__icontains="прожив") |
        Q(article__name__icontains="комнат")
    )
    rooms_ops = ops.filter(article__kind=DDSArticle.INCOME).filter(rooms_q)

    income_total = ops.filter(article__kind=DDSArticle.INCOME).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]
    expense_total = ops.filter(article__kind=DDSArticle.EXPENSE).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]
    balance = income_total - expense_total

    tz = timezone.get_current_timezone()
    rooms_by_day = (
        rooms_ops
        .annotate(day=TruncDate("happened_at", tzinfo=tz))
        .values("day")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("day")
    )

    wb = Workbook()

    # ── Лист 1: Итоги ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Итоги"
    label_font = Font(bold=True, size=11)
    for row_n, label in enumerate(["Отель", "Период", "", "Приход", "Расход", "Остаток"], 1):
        ws1.cell(row=row_n, column=1, value=label).font = label_font
    ws1.cell(row=1, column=2, value=hotel.name)
    ws1.cell(row=2, column=2, value=f"{date_from or '—'} → {date_to or '—'}")

    for row_n, val, fill_clr in [
        (4, float(income_total),  "E2EFDA"),
        (5, float(expense_total), "FCE4D6"),
        (6, float(balance),       "DDEEFF"),
    ]:
        fill = PatternFill("solid", fgColor=fill_clr)
        ws1.cell(row=row_n, column=1).fill = fill
        c = ws1.cell(row=row_n, column=2, value=val)
        c.number_format = _NUM_FMT
        c.alignment = Alignment(horizontal="right")
        c.border = _BORDER
        c.fill = fill

    _xl_widths(ws1, [("A", 14), ("B", 32)])

    # ── Лист 2: Номера по дням ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Номера по дням")
    _xl_hdr(ws2, 1, ["Дата", "Доход с номеров", "% от итого"])
    ws2.freeze_panes = "A2"

    rooms_list = list(rooms_by_day)
    rooms_total = sum((r["total"] for r in rooms_list), Decimal("0.00"))
    for i, r in enumerate(rooms_list, 2):
        _xl_row(ws2, i, [
            r["day"].strftime("%d.%m.%Y") if r["day"] else "",
            float(r["total"]),
            _pct(r["total"], rooms_total),
        ], money_cols=(2,), pct_cols=(3,))
    _xl_total_row(ws2, len(rooms_list) + 2,
                  ["ИТОГО", float(rooms_total), 1.0], money_cols=(2,), pct_cols=(3,))
    _xl_widths(ws2, [("A", 14), ("B", 20), ("C", 12)])

    # ── Лист 3: Операции ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Операции")
    _xl_hdr(ws3, 1, ["Дата", "Тип", "Статья", "Способ", "Сумма", "% от типа", "Контрагент", "Источник", "Комментарий"])
    ws3.freeze_panes = "A2"

    ops_list = list(ops.order_by("happened_at"))
    for i, op in enumerate(ops_list, 2):
        type_total = income_total if op.article.kind == DDSArticle.INCOME else expense_total
        _xl_row(ws3, i, [
            op.happened_at.strftime("%d.%m.%Y %H:%M"),
            op.article.get_kind_display(),
            op.article.name,
            op.get_method_display(),
            float(op.amount),
            _pct(op.amount, type_total),
            op.counterparty or "",
            op.source or "",
            (op.comment or "")[:500],
        ], money_cols=(5,), pct_cols=(6,))
    _xl_widths(ws3, [
        ("A", 16), ("B", 9), ("C", 28), ("D", 12),
        ("E", 14), ("F", 10), ("G", 22), ("H", 12), ("I", 35),
    ])

    df = date_from.strftime("%Y%m%d") if date_from else "start"
    dt = date_to.strftime("%Y%m%d") if date_to else "end"
    filename = f"hotel_{hotel.id}_dds_{df}_{dt}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def unified_report(request):
    # доступ: только superuser/finance_admin
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        # обычному пользователю можно показывать только его отельный дашборд
        return redirect("dds:dds_dashboard")

    hotels_qs = user_hotels_qs(request.user)  # для finance_admin будет "все"
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    ops = DDSOperation.objects.select_related("hotel", "article").filter(is_voided=False, hotel__in=hotels_qs)

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    # Свод по отелям
    by_hotels = (
        ops.values("hotel_id", "hotel__name")
        .annotate(
            income=Coalesce(Sum("amount", filter=Q(article__kind=DDSArticle.INCOME)), Decimal("0.00")),
            expense=Coalesce(Sum("amount", filter=Q(article__kind=DDSArticle.EXPENSE)), Decimal("0.00")),
        )
        .annotate(balance=F("income") - F("expense"))
        .order_by("hotel__name")
    )

    total_income = sum((x["income"] for x in by_hotels), Decimal("0.00"))
    total_expense = sum((x["expense"] for x in by_hotels), Decimal("0.00"))
    total_balance = total_income - total_expense

    # Свод по статьям (по сети)
    by_articles = (
        ops.values("article__kind", "article__name")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("article__kind", "-total")
    )

    return render(
        request,
        "dds/unified_report.html",
        {
            "by_hotels": by_hotels,
            "by_articles": by_articles,
            "total_income": total_income,
            "total_expense": total_expense,
            "total_balance": total_balance,
            "date_from": date_from,
            "date_to": date_to,
        },
    )



def _day_range(date_obj):
    start = timezone.make_aware(datetime.combine(date_obj, time.min))
    end = timezone.make_aware(datetime.combine(date_obj, time.max))
    return start, end


def _user_hotels_qs(user):
    # TODO: адаптируй под свою систему ролей
    if user.is_superuser or getattr(user, "is_finance_admin", False):
        return Hotel.objects.filter(is_active=True)
    # пример: профиль пользователя хранит отель
    hotel = getattr(getattr(user, "profile", None), "hotel", None)
    return Hotel.objects.filter(id=hotel.id, is_active=True) if hotel else Hotel.objects.none()



@login_required
def dds_dashboard(request):
    hotels_qs = _user_hotels_qs(request.user)

    hotel_id = request.GET.get("hotel") or ""
    selected_hotel = None
    if hotel_id:
        selected_hotel = get_object_or_404(hotels_qs, id=hotel_id)

    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    ops = (
        DDSOperation.objects
        .select_related("hotel", "article", "article__category", "article__category__parent")
        .filter(is_voided=False)
    )

    ops = ops.filter(hotel=selected_hotel) if selected_hotel else ops.filter(hotel__in=hotels_qs)

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    # =========================================================
    # Инкассация: НЕ показываем в графиках (и у тебя она ещё убрана из expense_sum)
    # =========================================================
    incasso_cut = Q(article__kind=DDSArticle.EXPENSE) & (
        Q(source__iexact="incasso") | Q(article__name__iexact="Инкассация")
    )
    ops_for_charts = ops.exclude(incasso_cut)

    # -----------------------------
    # Итоги
    # -----------------------------
    income_sum = ops.filter(article__kind=DDSArticle.INCOME).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    # ВАЖНО: сейчас расходы считаются БЕЗ инкассации (как ты и сделал)
    expense_sum = ops_for_charts.filter(article__kind=DDSArticle.EXPENSE).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    balance = income_sum - expense_sum

    # -----------------------------
    # Методы (счет)
    # -----------------------------
    methods = [DDSOperation.CASH, DDSOperation.MKASSA, DDSOperation.ZADATOK, DDSOperation.OPTIMA]
    method_labels = dict(DDSOperation.METHOD_CHOICES)
    method_headers = [{"code": m, "label": method_labels.get(m, m)} for m in methods]

    # =========================================================
    # 1) Таблицы: категории -> подкатегории + разбивка по методам
    # (таблицы строим по ops — инкассация там будет, если не хочешь, скажи)
    # =========================================================
    rows_qs = (
    ops_for_charts.values(
        "article__kind",
        "method",
        "article__category_id",
        "article__category__name",
        "article__category__parent_id",
        "article__category__parent__name",
    )
    .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
)

    rows = list(rows_qs)  # чтобы 2 раза не гонять один и тот же запрос

    def build_groups(kind: str):
        groups_map = {}
        uncat_by_method = defaultdict(lambda: Decimal("0.00"))

        for r in rows:
            if r["article__kind"] != kind:
                continue

            m = r["method"]
            total = r["total"] or Decimal("0.00")

            cat_id = r["article__category_id"]
            cat_name = r["article__category__name"]
            parent_id = r["article__category__parent_id"]
            parent_name = r["article__category__parent__name"]

            if not cat_id:
                uncat_by_method[m] += total
                continue

            if parent_id:
                group_id = parent_id
                group_name = parent_name or "Без названия"
                sub_id = cat_id
                sub_name = cat_name or "Без названия"
            else:
                group_id = cat_id
                group_name = cat_name or "Без названия"
                sub_id = None
                sub_name = None

            g = groups_map.get(group_id)
            if not g:
                g = {
                    "id": group_id,
                    "name": group_name,
                    "total": Decimal("0.00"),
                    "by_method": defaultdict(lambda: Decimal("0.00")),
                    "subs_map": {},
                }
                groups_map[group_id] = g

            g["total"] += total
            g["by_method"][m] += total

            if sub_id:
                s = g["subs_map"].get(sub_id)
                if not s:
                    s = {
                        "id": sub_id,
                        "name": sub_name,
                        "total": Decimal("0.00"),
                        "by_method": defaultdict(lambda: Decimal("0.00")),
                    }
                    g["subs_map"][sub_id] = s
                s["total"] += total
                s["by_method"][m] += total

        groups_list = []
        for g in groups_map.values():
            subs_list = sorted(g["subs_map"].values(), key=lambda x: (x["name"] or "").lower())
            for s in subs_list:
                s["method_totals"] = [s["by_method"][mm] for mm in methods]

            groups_list.append({
                "name": g["name"],
                "total": g["total"],
                "method_totals": [g["by_method"][mm] for mm in methods],
                "subs": subs_list,
            })

        groups_list.sort(key=lambda x: (x["name"] or "").lower())

        uncat = {
            "total": sum(uncat_by_method.values(), Decimal("0.00")),
            "method_totals": [uncat_by_method[mm] for mm in methods],
        }
        return groups_list, uncat

    income_groups, income_uncat = build_groups(DDSArticle.INCOME)
    expense_groups, expense_uncat = build_groups(DDSArticle.EXPENSE)

    # =========================================================
    # 2) Графики: по дням + stack по методам (по ops_for_charts)
    # =========================================================
    tz = timezone.get_current_timezone()
    day_rows_qs = (
        ops_for_charts
        .annotate(day=TruncDate("happened_at", tzinfo=tz))
        .values("day", "article__kind", "method")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("day")
    )
    day_rows = list(day_rows_qs)

    days = sorted({r["day"] for r in day_rows if r["day"] is not None})
    day_labels = [d.strftime("%Y-%m-%d") for d in days]

    grid = {
        DDSArticle.INCOME: {m: defaultdict(lambda: Decimal("0.00")) for m in methods},
        DDSArticle.EXPENSE: {m: defaultdict(lambda: Decimal("0.00")) for m in methods},
    }
    for r in day_rows:
        d = r["day"]
        if not d:
            continue
        kind = r["article__kind"]
        m = r["method"]
        if kind in grid and m in grid[kind]:
            grid[kind][m][d] += (r["total"] or Decimal("0.00"))

    def build_chart(kind: str):
        datasets = []
        for m in methods:
            datasets.append({
                "label": method_labels.get(m, m),
                "data": [float(grid[kind][m][d]) for d in days],
            })
        return {"labels": day_labels, "datasets": datasets}

    income_chart = build_chart(DDSArticle.INCOME)
    expense_chart = build_chart(DDSArticle.EXPENSE)

    # =========================================================
    # 3) Расходы: % по категориям (bar) + pie TOP N (+ Другое)
    #    тоже по ops_for_charts (инкассации там нет)
    # =========================================================
    expense_cat_rows_qs = (
        ops_for_charts.filter(article__kind=DDSArticle.EXPENSE)
        .values(
            "article__category_id",
            "article__category__name",
            "article__category__parent_id",
            "article__category__parent__name",
        )
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
    )
    expense_cat_rows = list(expense_cat_rows_qs)

    cat_totals = defaultdict(lambda: Decimal("0.00"))
    for r in expense_cat_rows:
        total = r["total"] or Decimal("0.00")

        if not r["article__category_id"]:
            group_name = "Без категории"
        else:
            parent_id = r["article__category__parent_id"]
            parent_name = r["article__category__parent__name"]
            cat_name = r["article__category__name"]
            group_name = (parent_name if parent_id else cat_name) or "Без категории"

        cat_totals[group_name] += total

    cat_sorted = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)
    grand_total = sum(cat_totals.values(), Decimal("0.00"))




    expense_cat_percent = {"labels": [], "percent": [], "amounts": [], "grand_total": 0.0}

    if grand_total > 0:
        expense_cat_percent["grand_total"] = float(grand_total)

        for name, total in cat_sorted:
        # точный процент
            pct = (total / grand_total) * Decimal("100")
            pct = pct.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            expense_cat_percent["labels"].append(name)
            expense_cat_percent["percent"].append(float(pct))
            expense_cat_percent["amounts"].append(float(total))

    

    TOP_N = 8
    pie_labels, pie_data = [], []
    other_sum = Decimal("0.00")

    for i, (name, total) in enumerate(cat_sorted):
        if i < TOP_N:
            pie_labels.append(name)
            pie_data.append(float(total))
        else:
            other_sum += total

    if other_sum > 0:
        pie_labels.append("Другое")
        pie_data.append(float(other_sum))

    expense_cat_share = {
        "labels": pie_labels,
        "data": pie_data,
        "grand_total": float(grand_total),  # важно для tooltip в твоём JS
    }
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:dds_dashboard")

    hotels = user_hotels_qs(request.user).order_by("name")

    hotel_id = request.GET.get("hotel")
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    # фильтр по отелю
    hotels_filter = hotels
    if hotel_id:
        hotels_filter = hotels.filter(id=hotel_id)

    

    # 2) ИНКАССАЦИИ
    incassos = CashIncasso.objects.select_related("hotel").filter(hotel__in=hotels_filter)
    return render(request, "dds/dashboard.html", {
        "hotels": hotels_qs,
        "selected_hotel": selected_hotel,
        "date_from": date_from,
        "date_to": date_to,
        "incassos": incassos.order_by("-happened_at")[:300], 
        "income_sum": income_sum,
        "expense_sum": expense_sum,
        "balance": balance,

        "method_headers": method_headers,

        "income_groups": income_groups,
        "income_uncat": income_uncat,
        "expense_groups": expense_groups,
        "expense_uncat": expense_uncat,

        "income_chart": income_chart,
        "expense_chart": expense_chart,

        "expense_cat_percent": expense_cat_percent,
        "expense_cat_share": expense_cat_share,
    })


@login_required
def dds_list(request):
    hotels_qs = _user_hotels_qs(request.user)

    ops = DDSOperation.objects.select_related("hotel", "article").filter(hotel__in=hotels_qs)

    hotel_id = request.GET.get("hotel")
    kind = request.GET.get("kind")  # income/expense
    article_id = request.GET.get("article")
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    if hotel_id:
        ops = ops.filter(hotel_id=hotel_id)
    if kind in (DDSArticle.INCOME, DDSArticle.EXPENSE):
        ops = ops.filter(article__kind=kind)
    if article_id:
        ops = ops.filter(article_id=article_id)
    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    articles = DDSArticle.objects.filter(is_active=True)

    return render(
        request,
        "dds/operation_list.html",
        {
            "ops": ops[:500],  # на MVP
            "hotels": hotels_qs,
            "articles": articles,
            "filters": {"hotel": hotel_id, "kind": kind, "article": article_id, "date_from": date_from, "date_to": date_to},
        },
    )




def _is_rooms_income(op) -> bool:
    src = (op.source or "").lower()
    name = (op.article.name or "").lower()
    return (
        src == "rooms"
        or "room" in src
        or ("номер" in name)
        or ("прожив" in name)
        or ("комнат" in name)
    )


@login_required
def dds_op_add(request, hotel_id: int, kind: str):
    hotels = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels, id=hotel_id)

    register, _ = CashRegister.objects.get_or_create(hotel=hotel)

    category_id = request.GET.get("category") or request.POST.get("category") or None

    if request.method == "POST":
        form = DDSOpCreateForm(
            request.POST,
            kind=kind,
            category_id=category_id,
            hotel=hotel,   # ✅ ВАЖНО: без этого будет показывать всё
        )

        if form.is_valid():
            op = form.save(commit=False)
            op.hotel = hotel
            op.created_by = request.user

            direction = CashMovement.IN if op.article.kind == DDSArticle.INCOME else CashMovement.OUT
            is_incasso = (op.source or "").lower() == "incasso"

            try:
                with transaction.atomic():
                    register = CashRegister.objects.select_for_update().get(pk=register.pk)

                    if (not is_incasso) and direction == CashMovement.OUT:
                        field = FIELD_MAP.get(op.method)
                        if not field:
                            messages.error(request, "Неверный способ оплаты/счет.")
                            return render(request, "dds/dds_quick_op_form.html", {
                                "hotel": hotel, "register": register, "kind": kind,
                                "form": form, "category_id": category_id or "",
                            })

                        current = getattr(register, field) or Decimal("0.00")
                        if op.amount > current:
                            messages.error(
                                request,
                                f"Недостаточно средств на счете {op.get_method_display()}. Доступно: {current}"
                            )
                            return render(request, "dds/dds_quick_op_form.html", {
                                "hotel": hotel, "register": register, "kind": kind,
                                "form": form, "category_id": category_id or "",
                            })

                    op.save()

                    if not is_incasso:
                        exists = CashMovement.objects.filter(
                            dds_operation=op,
                            account=op.method,
                            direction=direction,
                        ).exists()

                        if not exists:
                            try:
                                apply_cash_movement(
                                    hotel=hotel,
                                    account=op.method,
                                    direction=direction,
                                    amount=op.amount,
                                    created_by=request.user,
                                    happened_at=op.happened_at,
                                    comment=op.comment,
                                    dds_operation=op,
                                )
                            except IntegrityError:
                                pass

            except ValidationError as e:
                messages.error(request, str(e))
                return render(request, "dds/dds_quick_op_form.html", {
                    "hotel": hotel, "register": register, "kind": kind,
                    "form": form, "category_id": category_id or "",
                })

            messages.success(request, "Операция сохранена и касса обновлена.")
            return redirect("dds:hotel_detail", hotel.id)

        messages.error(request, "Исправьте ошибки в форме.")

    else:
        form = DDSOpCreateForm(
            kind=kind,
            category_id=category_id,
            hotel=hotel,   # ✅ ВАЖНО
        )

    return render(request, "dds/dds_quick_op_form.html", {
        "hotel": hotel,
        "register": register,
        "kind": kind,
        "form": form,
        "category_id": category_id or "",
    })

@login_required
def dds_create(request):
    hotels_qs = user_hotels_qs(request.user)
    if not hotels_qs.exists():
        messages.error(request, "У вас не назначен отель. Обратитесь к администратору.")
        return redirect("dds:dds_list")

    only_hotel = hotels_qs.first() if hotels_qs.count() == 1 else None

    selected_hotel = None
    if only_hotel:
        selected_hotel = only_hotel
    else:
        hotel_id = request.POST.get("hotel") if request.method == "POST" else request.GET.get("hotel")
        if hotel_id and hotels_qs.filter(id=hotel_id).exists():
            selected_hotel = hotels_qs.get(id=hotel_id)

    kind = request.GET.get("kind") or request.POST.get("kind")
    if kind not in (DDSArticle.INCOME, DDSArticle.EXPENSE):
        kind = None

    if request.method == "POST":
        form = DDSOperationForm(request.POST, hotel=selected_hotel, kind=kind)
        form.fields["hotel"].queryset = hotels_qs


        if only_hotel:
            form.fields["hotel"].initial = only_hotel
            form.fields["hotel"].disabled = True

        # ❌ НЕ ДЕЛАЙ ТАК:
        # form.fields["article"].queryset = DDSArticle.objects.filter(is_active=True)

        if form.is_valid():
            op = form.save(commit=False)
            
            
            if only_hotel:
            
                op.hotel = only_hotel
            op.created_by = request.user
            op.save()

            messages.success(request, "Операция добавлена.")
            return redirect("dds:hotel_detail", pk=op.hotel_id)

    else:
        form = DDSOperationForm(hotel=selected_hotel, kind=kind)



        form.fields["hotel"].queryset = hotels_qs




        if selected_hotel:
            form.fields["hotel"].initial = selected_hotel
        if only_hotel:
            form.fields["hotel"].initial = only_hotel
            form.fields["hotel"].disabled = True

    
    reg = None
    
    if selected_hotel:
        reg, _ = CashRegister.objects.get_or_create(hotel=selected_hotel)

    return render(request, "dds/operation_form.html", {
        "form": form,
        "reg": reg,
        "selected_hotel": selected_hotel,
        "kind": kind,
    })





@login_required
def dds_void(request, pk):
    hotels_qs = _user_hotels_qs(request.user)
    op = get_object_or_404(DDSOperation, pk=pk, hotel__in=hotels_qs)

    if request.method == "POST":
        reason = (request.POST.get("reason") or "").strip()
        if not reason:
            messages.error(request, "Укажи причину сторно.")
            return redirect("dds:dds_list")
        op.void(request.user, reason)
        messages.success(request, "Операция отменена (сторно).")
        return redirect("dds:dds_list")

    return render(request, "dds/void_confirm.html", {"op": op})


@login_required
def dds_articles(request):
    # TODO: доступ только бухгалтеру/админу
    if not (request.user.is_superuser or getattr(request.user, "is_finance_admin", False)):
        return redirect("dds:dds_dashboard")

    if request.method == "POST":
        form = DDSArticleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Статья сохранена.")
            return redirect("dds:dds_articles")
    else:
        form = DDSArticleForm()

    articles = DDSArticle.objects.all()
    return render(request, "dds/articles.html", {"form": form, "articles": articles})





@login_required
def hotel_catalog(request):
    # каталог менять/добавлять — только админ/финанс
    is_fin_admin = request.user.is_superuser or getattr(getattr(request.user, "profile", None), "is_finance_admin", False)

    # что показываем:
    hotels = Hotel.objects.all().order_by("name") if is_fin_admin else user_hotels_qs(request.user).order_by("name")

    form = None
    if is_fin_admin:
        form = HotelForm(request.POST or None)
        if request.method == "POST" and form.is_valid():
            form.save()
            messages.success(request, "Отель добавлен.")
            return redirect("dds:hotel_catalog")

    return render(request, "dds/hotel_catalog.html", {"hotels": hotels, "form": form, "is_fin_admin": is_fin_admin})




@login_required
def hotel_list(request):
    hotels = list(user_hotels_qs(request.user).order_by("name"))

    # Берём кассы одним запросом и мапим по hotel_id
    registers_by_hotel = CashRegister.objects.filter(hotel__in=hotels).in_bulk(field_name="hotel_id")

    # приклеиваем register к каждому отелю (может быть None)
    for h in hotels:
        h.register = registers_by_hotel.get(h.id)

    return render(request, "dds/hotel_list.html", {"hotels": hotels})


@login_required
def hotel_detail(request, pk):
    hotels_qs = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels_qs, pk=pk)

    reg, _ = CashRegister.objects.get_or_create(hotel=hotel)

    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    ops = (
        DDSOperation.objects
        .select_related(
            "article",
            "article__category",
            "article__category__parent",
        )
        .filter(hotel=hotel, is_voided=False)
    )

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    # ✅ Итоги ДДС за период
    income_total = ops.filter(article__kind=DDSArticle.INCOME).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    expense_total = ops.filter(article__kind=DDSArticle.EXPENSE).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    balance = income_total - expense_total

    # ✅ Доход с номеров
    rooms_q = (
        Q(source__iexact="rooms") |
        Q(source__icontains="room") |
        Q(article__name__icontains="номер") |
        Q(article__name__icontains="прожив") |
        Q(article__name__icontains="комнат")
    )
    rooms_ops = ops.filter(article__kind=DDSArticle.INCOME).filter(rooms_q)

    rooms_income_total = rooms_ops.aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    tz = timezone.get_current_timezone()
    rooms_by_day = (
        rooms_ops
        .annotate(day=TruncDate("happened_at", tzinfo=tz))
        .values("day")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("day")
    )

    last_ops = ops.order_by("-happened_at")[:50]

    # ✅ ДДС по счетам за период (как у тебя)
    methods = [DDSOperation.CASH, DDSOperation.MKASSA, DDSOperation.ZADATOK, DDSOperation.OPTIMA]
    method_labels = dict(DDSOperation.METHOD_CHOICES)

    period_rows = []
    for m in methods:
        inc = ops.filter(article__kind=DDSArticle.INCOME, method=m).aggregate(
            s=Coalesce(Sum("amount"), Decimal("0.00"))
        )["s"]
        exp = ops.filter(article__kind=DDSArticle.EXPENSE, method=m).aggregate(
            s=Coalesce(Sum("amount"), Decimal("0.00"))
        )["s"]
        period_rows.append({
            "code": m,
            "label": method_labels.get(m, m),
            "income": inc,
            "expense": exp,
            "delta": inc - exp,
        })

    # ==========================================================
    # ✅ НОВОЕ: Доходы/Расходы по категориям и подкатегориям
    # ==========================================================
    def build_cat_groups(kind: str):
        """
        Возвращает:
          groups: [
            {id, name, total, subs:[{id,name,total}, ...]},
            ...
          ]
          uncategorized_total: Decimal
        """
        rows = (
            ops.filter(article__kind=kind)
            .values(
                "article__category_id",
                "article__category__name",
                "article__category__parent_id",
                "article__category__parent__name",
            )
            .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        )

        # group_id = parent_id если есть, иначе category_id (верхний уровень)
        groups = {}  # group_id -> dict
        uncategorized_total = Decimal("0.00")

        for r in rows:
            cat_id = r["article__category_id"]
            cat_name = r["article__category__name"]
            parent_id = r["article__category__parent_id"]
            parent_name = r["article__category__parent__name"]
            total = r["total"] or Decimal("0.00")

            if not cat_id:
                uncategorized_total += total
                continue

            if parent_id:
                # подкатегория -> складываем в родителя
                group_id = parent_id
                group_name = parent_name or "Без названия"
                g = groups.get(group_id)
                if not g:
                    g = {"id": group_id, "name": group_name, "total": Decimal("0.00"), "subs": []}
                    groups[group_id] = g

                g["subs"].append({"id": cat_id, "name": cat_name or "Без названия", "total": total})
                g["total"] += total
            else:
                # верхняя категория
                group_id = cat_id
                group_name = cat_name or "Без названия"
                g = groups.get(group_id)
                if not g:
                    g = {"id": group_id, "name": group_name, "total": Decimal("0.00"), "subs": []}
                    groups[group_id] = g
                # если операции записаны прямо на верхнюю категорию — добавляем в total
                g["total"] += total

        # сортировка
        groups_list = sorted(groups.values(), key=lambda x: (x["name"] or "").lower())
        for g in groups_list:
            g["subs"] = sorted(g["subs"], key=lambda x: (x["name"] or "").lower())

        return groups_list, uncategorized_total

    income_groups, income_uncat = build_cat_groups(DDSArticle.INCOME)
    expense_groups, expense_uncat = build_cat_groups(DDSArticle.EXPENSE)

    return render(request, "dds/hotel_detail.html", {
        "hotel": hotel,
        "reg": reg,

        # фильтр периода
        "date_from": date_from,
        "date_to": date_to,

        # итоги
        "income_total": income_total,
        "expense_total": expense_total,
        "balance": balance,

        # номера
        "rooms_income_total": rooms_income_total,
        "rooms_by_day": rooms_by_day,

        # по счетам за период
        "period_rows": period_rows,

        # последние операции
        "last_ops": last_ops,

        # ✅ категории/подкатегории
        "income_groups": income_groups,
        "income_uncat": income_uncat,
        "expense_groups": expense_groups,
        "expense_uncat": expense_uncat,
    })




@login_required
def unified_report_export_excel(request):
    # ✅ доступ только superuser/finance_admin
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:dds_dashboard")

    hotels_qs = user_hotels_qs(request.user)

    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    ops = DDSOperation.objects.select_related("hotel", "article").filter(
        is_voided=False,
        hotel__in=hotels_qs,
    )

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    # ✅ Свод по отелям
    by_hotels = (
        ops.values("hotel_id", "hotel__name")
        .annotate(
            income=Coalesce(Sum("amount", filter=Q(article__kind=DDSArticle.INCOME)), Decimal("0.00")),
            expense=Coalesce(Sum("amount", filter=Q(article__kind=DDSArticle.EXPENSE)), Decimal("0.00")),
        )
        .annotate(balance=F("income") - F("expense"))
        .order_by("hotel__name")
    )

    total_income = sum((x["income"] for x in by_hotels), Decimal("0.00"))
    total_expense = sum((x["expense"] for x in by_hotels), Decimal("0.00"))
    total_balance = total_income - total_expense

    # ✅ Свод по статьям (по сети)
    by_articles = (
        ops.values("article__kind", "article__name")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("article__kind", "-total", "article__name")
    )

    # ── Excel ──────────────────────────────────────────────────────────────
    wb = Workbook()
    label_font = Font(bold=True, size=11)

    # Лист 1: Итоги
    ws1 = wb.active
    ws1.title = "Итоги"
    ws1.cell(row=1, column=1, value="Отчет").font  = label_font
    ws1.cell(row=1, column=2, value="Единый отчет по сети (ДДС)")
    ws1.cell(row=2, column=1, value="Период").font = label_font
    ws1.cell(row=2, column=2, value=f"{date_from or '—'} → {date_to or '—'}")

    for row_n, label, val, fill_clr in [
        (4, "Итого приход",  float(total_income),  "E2EFDA"),
        (5, "Итого расход",  float(total_expense), "FCE4D6"),
        (6, "Итого остаток", float(total_balance), "DDEEFF"),
    ]:
        fill = PatternFill("solid", fgColor=fill_clr)
        lc = ws1.cell(row=row_n, column=1, value=label)
        lc.font = label_font; lc.fill = fill
        vc = ws1.cell(row=row_n, column=2, value=val)
        vc.number_format = _NUM_FMT
        vc.alignment = Alignment(horizontal="right")
        vc.border = _BORDER
        vc.fill = fill

    _xl_widths(ws1, [("A", 18), ("B", 40)])

    # Лист 2: По отелям
    ws2 = wb.create_sheet("По отелям")
    _xl_hdr(ws2, 1, ["Отель", "Приход", "% прих.", "Расход", "% расх.", "Остаток"])
    ws2.freeze_panes = "A2"

    hotels_list = list(by_hotels)
    for i, r in enumerate(hotels_list, 2):
        _xl_row(ws2, i, [
            r["hotel__name"],
            float(r["income"]),
            _pct(r["income"], total_income),
            float(r["expense"]),
            _pct(r["expense"], total_expense),
            float(r["balance"]),
        ], money_cols=(2, 4, 6), pct_cols=(3, 5))
    _xl_total_row(ws2, len(hotels_list) + 2,
                  ["ИТОГО", float(total_income), 1.0, float(total_expense), 1.0, float(total_balance)],
                  money_cols=(2, 4, 6), pct_cols=(3, 5))
    _xl_widths(ws2, [("A", 32), ("B", 16), ("C", 10), ("D", 16), ("E", 10), ("F", 16)])

    # Лист 3: По статьям
    ws3 = wb.create_sheet("По статьям")
    _xl_hdr(ws3, 1, ["Тип", "Статья", "Сумма", "% от типа"])
    ws3.freeze_panes = "A2"

    articles_list = list(by_articles)
    inc_total_art = sum(
        (r["total"] for r in articles_list if r["article__kind"] == DDSArticle.INCOME),
        Decimal("0.00"),
    )
    exp_total_art = sum(
        (r["total"] for r in articles_list if r["article__kind"] != DDSArticle.INCOME),
        Decimal("0.00"),
    )
    for i, r in enumerate(articles_list, 2):
        kind_label = "Доход" if r["article__kind"] == DDSArticle.INCOME else "Расход"
        type_total = inc_total_art if r["article__kind"] == DDSArticle.INCOME else exp_total_art
        _xl_row(ws3, i, [
            kind_label, r["article__name"], float(r["total"]), _pct(r["total"], type_total),
        ], money_cols=(3,), pct_cols=(4,))
    last = len(articles_list) + 2
    _xl_total_row(ws3, last,     ["Доход",  "ИТОГО доход",  float(inc_total_art), 1.0], money_cols=(3,), pct_cols=(4,))
    _xl_total_row(ws3, last + 1, ["Расход", "ИТОГО расход", float(exp_total_art), 1.0], money_cols=(3,), pct_cols=(4,))
    _xl_widths(ws3, [("A", 10), ("B", 38), ("C", 16), ("D", 12)])

    df = date_from.strftime("%Y%m%d") if date_from else "start"
    dt = date_to.strftime("%Y%m%d") if date_to else "end"
    filename = f"unified_report_{df}_{dt}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def incasso_create(request, pk):
    hotels_qs = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels_qs, pk=pk)

    reg, _ = CashRegister.objects.get_or_create(hotel=hotel)

    if request.method == "POST":
        form = CashIncassoForm(request.POST)
        if form.is_valid():
            inc = form.save(commit=False)
            inc.hotel = hotel
            inc.created_by = request.user

            try:
                with transaction.atomic():
                    reg = CashRegister.objects.select_for_update().get(pk=reg.pk)

                    field = FIELD_MAP.get(inc.method)
                    current = getattr(reg, field) if field else None
                    current = current or Decimal("0.00")

                    if inc.amount > current:
                        messages.error(
                            request,
                            f"Недостаточно средств на счете {inc.get_method_display()}. Доступно: {current}"
                        )
                        return render(request, "dds/incasso_form.html", {"form": form, "hotel": hotel, "reg": reg})

                    inc.save()

                    article, _ = DDSArticle.objects.get_or_create(
                        kind=DDSArticle.EXPENSE,
                        name="Инкассация",
                        defaults={"is_active": True},
                    )

                    op_incasso = DDSOperation.objects.create(
                        hotel=hotel,
                        article=article,
                        amount=inc.amount,
                        happened_at=inc.happened_at,
                        method=inc.method,      # ✅ с какого счета забрали
                        counterparty="Бухгалтерия",
                        comment=inc.comment,
                        source="incasso",
                        created_by=request.user,
                    )

                    apply_cash_movement(
                        hotel=hotel,
                        account=inc.method,         # ✅ списываем именно отсюда
                        direction=CashMovement.OUT,
                        amount=inc.amount,
                        created_by=request.user,
                        happened_at=inc.happened_at,
                        comment=inc.comment,
                        dds_operation=op_incasso,
                        incasso=inc,
                    )

            except ValidationError as e:
                messages.error(request, str(e))
                return render(request, "dds/incasso_form.html", {"form": form, "hotel": hotel, "reg": reg})

            messages.success(request, "Инкассация создана. Средства списаны.")
            return redirect("dds:hotel_detail", pk=hotel.id)
    else:
        form = CashIncassoForm()

    return render(request, "dds/incasso_form.html", {"form": form, "hotel": hotel, "reg": reg})


@login_required
def accounting(request):
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:dds_dashboard")

    hotels = user_hotels_qs(request.user).order_by("name")

    hotel_id = request.GET.get("hotel")
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    # фильтр по отелю
    hotels_filter = hotels
    if hotel_id:
        hotels_filter = hotels.filter(id=hotel_id)

    # 1) РАСХОДЫ (не инкассация!)
    expenses = DDSOperation.objects.select_related("hotel", "article").filter(
        is_voided=False,
        hotel__in=hotels_filter,
        article__kind=DDSArticle.EXPENSE,
    ).exclude(source="incasso")

    # 2) ИНКАССАЦИИ
    incassos = CashIncasso.objects.select_related("hotel").filter(hotel__in=hotels_filter)

    # фильтр по датам
    if date_from:
        start, _ = _day_range(date_from)
        expenses = expenses.filter(happened_at__gte=start)
        incassos = incassos.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        expenses = expenses.filter(happened_at__lte=end)
        incassos = incassos.filter(happened_at__lte=end)

    expense_total = expenses.aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]
    incasso_total = incassos.aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]

    return render(request, "dds/accounting.html", {
        "hotels": hotels,
        "selected_hotel": hotel_id,
        "date_from": date_from,
        "date_to": date_to,
        "expenses": expenses.order_by("-happened_at")[:300],
        "incassos": incassos.order_by("-happened_at")[:300],
        "expense_total": expense_total,
        "incasso_total": incasso_total,
    })


@login_required
def accounting_export_excel(request):
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:dds_dashboard")

    hotels = user_hotels_qs(request.user)

    hotel_id = request.GET.get("hotel")
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    hotels_filter = hotels
    if hotel_id:
        hotels_filter = hotels.filter(id=hotel_id)

    expenses = DDSOperation.objects.select_related("hotel", "article").filter(
        is_voided=False,
        hotel__in=hotels_filter,
        article__kind=DDSArticle.EXPENSE,
    ).exclude(source="incasso")

    incassos = CashIncasso.objects.select_related("hotel").filter(hotel__in=hotels_filter)

    if date_from:
        start, _ = _day_range(date_from)
        expenses = expenses.filter(happened_at__gte=start)
        incassos = incassos.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        expenses = expenses.filter(happened_at__lte=end)
        incassos = incassos.filter(happened_at__lte=end)

    exp_total = expenses.aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]
    inc_total = incassos.aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]

    wb = Workbook()
    label_font = Font(bold=True, size=11)

    # ── Лист 1: Итоги ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Итоги"
    ws1.cell(row=1, column=1, value="Период").font    = label_font
    ws1.cell(row=1, column=2, value=f"{date_from or '—'} → {date_to or '—'}")
    ws1.cell(row=2, column=1, value="Отель").font     = label_font
    ws1.cell(row=2, column=2, value=hotel_id or "Все")

    for row_n, label, val, fill_clr in [
        (4, "Расходы (без инкассации)", float(exp_total), "FCE4D6"),
        (5, "Инкассации",               float(inc_total), "FFF2CC"),
    ]:
        fill = PatternFill("solid", fgColor=fill_clr)
        lc = ws1.cell(row=row_n, column=1, value=label)
        lc.font = label_font; lc.fill = fill
        vc = ws1.cell(row=row_n, column=2, value=val)
        vc.number_format = _NUM_FMT
        vc.alignment = Alignment(horizontal="right")
        vc.border = _BORDER
        vc.fill = fill

    _xl_widths(ws1, [("A", 28), ("B", 32)])

    # ── Лист 2: Расходы ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Расходы")
    _xl_hdr(ws2, 1, ["Дата", "Отель", "Статья", "Способ", "Сумма", "% от итого", "Контрагент", "Комментарий"])
    ws2.freeze_panes = "A2"

    expenses_list = list(expenses.order_by("happened_at"))
    for i, op in enumerate(expenses_list, 2):
        _xl_row(ws2, i, [
            op.happened_at.strftime("%d.%m.%Y %H:%M"),
            op.hotel.name,
            op.article.name,
            op.get_method_display(),
            float(op.amount),
            _pct(op.amount, exp_total),
            op.counterparty or "",
            (op.comment or "")[:500],
        ], money_cols=(5,), pct_cols=(6,))
    _xl_total_row(ws2, len(expenses_list) + 2,
                  ["", "", "", "ИТОГО", float(exp_total), 1.0, "", ""],
                  money_cols=(5,), pct_cols=(6,))
    _xl_widths(ws2, [
        ("A", 16), ("B", 22), ("C", 28), ("D", 12), ("E", 14), ("F", 10), ("G", 22), ("H", 35),
    ])

    # ── Лист 3: Инкассации ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Инкассации")
    _xl_hdr(ws3, 1, ["Дата", "Отель", "Способ", "Сумма", "% от итого", "Комментарий", "Создал"])
    ws3.freeze_panes = "A2"

    incasso_list = list(incassos.order_by("happened_at"))
    for i, inc in enumerate(incasso_list, 2):
        _xl_row(ws3, i, [
            inc.happened_at.strftime("%d.%m.%Y %H:%M"),
            inc.hotel.name,
            inc.get_method_display(),
            float(inc.amount),
            _pct(inc.amount, inc_total),
            (inc.comment or "")[:500],
            getattr(inc.created_by, "username", ""),
        ], money_cols=(4,), pct_cols=(5,))
    _xl_total_row(ws3, len(incasso_list) + 2,
                  ["", "", "ИТОГО", float(inc_total), 1.0, "", ""],
                  money_cols=(4,), pct_cols=(5,))
    _xl_widths(ws3, [("A", 16), ("B", 22), ("C", 12), ("D", 14), ("E", 10), ("F", 35), ("G", 16)])

    df = date_from.strftime("%Y%m%d") if date_from else "start"
    dt = date_to.strftime("%Y%m%d") if date_to else "end"
    filename = f"accounting_{df}_{dt}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

